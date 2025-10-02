from openpyxl import load_workbook

# Mapeamento das colunas do Excel
COLUNAS = {
    "JAN": 2, "FEV": 3, "MAR": 4, "ABR": 5, "MAI": 6, "JUN": 7,
    "JUL": 8, "AGO": 9, "SET": 10, "OUT": 11, "NOV": 12, "DEZ": 13
}

def atualizar_receita_bruta(caminho_xlsx: str, mes: str, valor: float, saida: str):
    """
    Atualiza a linha 'Receita Operacional Bruta' na planilha DRE.
    Gera um novo arquivo com nome definido em 'saida'.
    """
    wb = load_workbook(caminho_xlsx)
    ws = wb.active

    linha_receita = None
    for row in ws.iter_rows(min_row=1, max_col=1):
        if row[0].value == "Receita Operacional Bruta":
            linha_receita = row[0].row
            break

    if not linha_receita:
        raise ValueError("Linha 'Receita Operacional Bruta' não encontrada na planilha!")

    if mes not in COLUNAS:
        raise ValueError(f"Mês inválido: {mes}. Use JAN, FEV, MAR...")

    ws.cell(row=linha_receita, column=COLUNAS[mes], value=valor)
    wb.save(saida)
    print(f"✅ Planilha atualizada: {saida}")

